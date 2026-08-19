import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from scripts.payroll_calculator import attendance_records, calculate, read_rows, write_workbook


class PayrollCalculatorTest(unittest.TestCase):
    def make_book(self, directory, name, headers, rows):
        path = Path(directory) / name
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        workbook.save(path)
        return path

    def test_aliases_partial_baseline_and_rules(self):
        with tempfile.TemporaryDirectory() as directory:
            roster = self.make_book(
                directory, "roster.xlsx", ["姓名", "项目", "人员类别", "状态"],
                [["张三", "A项目", "外聘", "正常"], ["李四", "A项目", "外聘", "在册"]],
            )
            history = self.make_book(
                directory, "history.xlsx",
                ["姓名", "项目", "人员类别", "基本工资", "绩效工资", "工龄工资",
                 "职称工资", "施工补贴", "1-3月话补", "降温费", "交通补助", "应发工资"],
                [["张三", "A项目", "外聘", 10000, 3000, 200, 100, 50, 300, 200, 100, 0],
                 ["李四", "A项目", "外聘", 9000, 2000, 0, 0, 40, 200, 0, 0, 0]],
            )
            rows = calculate(
                read_rows(roster), read_rows(history),
                [{"name": "张三", "project": "A项目", "category": "外聘",
                  "actual_work_days": 23, "personal_leave_days": 2}],
                "2026-06", 21.75, 2,
            )
            detail = rows["payroll_detail"][0]
            self.assertEqual(detail["绩效工资"], 3000 - 3000 / 21.75 * 2)
            self.assertEqual(detail["施工补贴"], 0)
            self.assertEqual(detail["加班天数"], 1)
            self.assertEqual(rows["attendance"][0]["来源"], "当月考勤表")
            self.assertTrue(any("施工补贴只有月金额" in item["说明"]
                                for item in rows["review_exceptions"]))
            self.assertTrue(any(item["类型"] == "可忽略" for item in rows["review_exceptions"]))
            self.assertTrue(any(item.get("姓名") == "张三" for item in rows["review_exceptions"]))

            output = Path(directory) / "result.xlsx"
            write_workbook(rows, output)
            workbook = load_workbook(output, read_only=True)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["工资核算明细", "工资基准", "考勤汇总", "历史工资对比", "人工复核事项"],
                )
            finally:
                workbook.close()

    def test_missing_workdays_does_not_block_other_items(self):
        rows = calculate(
            [], [{"name": "王五", "project": "B", "category": "外聘",
                  "position_salary": 5000, "performance": 1000}],
            [{"name": "王五", "project": "B", "category": "外聘",
              "actual_work_days": 20}],
            "2026-05", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["岗位工资"], 5000)
        self.assertEqual(rows["payroll_detail"][0]["加班费"], 0)
        self.assertEqual(rows["payroll_detail"][0]["加班天数"], 0)

    def test_missing_attendance_is_not_written_as_zero(self):
        rows = calculate(
            [], [{"name": "缺考勤", "project": "B", "category": "外聘",
                  "position_salary": 5000, "performance": 1000}],
            [{"name": "缺考勤", "project": "B", "category": "外聘"}],
            "2026-06", 21, 4,
        )
        detail = rows["payroll_detail"][0]
        self.assertIsNone(detail["实际出勤"])
        self.assertTrue(any("缺少实际出勤" in item["说明"] for item in rows["review_exceptions"]))

    def test_construction_total_is_normalized_to_daily_standard(self):
        rows = calculate(
            [],
            [{"name": "赵六", "项目": "C", "category": "外聘",
              "work_days": 26, "position_salary": 5000,
              "performance": 1000, "construction": 780}],
            [{"name": "赵六", "project": "C", "category": "外聘",
              "actual_work_days": 23}],
            "2026-06", 22, 8,
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["施工补贴"], 690)

    def test_hot_subsidy_requires_individual_history_and_explicit_overtime_is_counted(self):
        rows = calculate(
            [],
            [{"name": "甲", "项目": "D", "category": "外聘",
              "position_salary": 5000, "performance": 1000,
              "hot": 600, "construction_day": 30},
             {"name": "乙", "项目": "D", "category": "外聘",
              "position_salary": 5000, "performance": 1000,
              "hot": 0, "construction_day": 30}],
            [{"name": "甲", "project": "D", "category": "外聘",
              "actual_work_days": 22, "加班天数": 2},
             {"name": "乙", "project": "D", "category": "外聘",
              "actual_work_days": 22}],
            "2026-06", 22, 8,
        )
        details = {row["姓名"]: row for row in rows["payroll_detail"]}
        baselines = {row["姓名"]: row for row in rows["baseline"]}
        self.assertEqual(baselines["甲"]["高温补贴"], 600)
        self.assertEqual(baselines["乙"]["高温补贴"], 0)
        self.assertEqual(details["甲"]["加班天数"], 0)
        self.assertEqual(details["甲"]["加班费"], 0)

    def test_quarter_phone_requires_manual_confirmation(self):
        rows = calculate(
            [],
            [{"name": "钱七", "项目": "E", "category": "外聘",
              "position_salary": 5000, "performance": 1000,
              "phone_1_3": 300}],
            [{"name": "钱七", "project": "E", "category": "外聘",
              "actual_work_days": 22}],
            "2026-03", 22, 0,
        )
        self.assertEqual(rows["payroll_detail"][0]["话费补贴"], 0)
        self.assertTrue(any("季度话补" in item["说明"] for item in rows["review_exceptions"]))

    def test_overtime_counts_attendance_on_single_rest_day(self):
        rows = calculate(
            [],
            [{"name": "孙八", "项目": "F", "category": "外聘",
              "position_salary": 5000, "performance": 1000}],
            [{"name": "孙八", "project": "F", "category": "外聘",
              "actual_work_days": 27,
              "daily_marks": {
                  "2026-06-06": "出勤",  # Saturday is a planned workday for single rest.
                  "2026-06-07": "出勤",  # Both weekend days worked: two overtime days.
              }}],
            "2026-06", 26, 4, "single",
        )
        self.assertEqual(rows["payroll_detail"][0]["加班天数"], 4)

    def test_overtime_counts_one_day_for_each_worked_weekend_without_schedule(self):
        rows = calculate(
            [], [{"name": "周末", "project": "瀚阅府", "category": "益通外包",
                  "position_salary": 3800, "performance": 1000}],
            [{"name": "周末", "project": "瀚阅府", "category": "益通外包",
              "actual_work_days": 24, "weekend_attendance_dates": [
                  "2026-06-06", "2026-06-14", "2026-06-20", "2026-06-28",
              ]}],
            "2026-06", None, None,
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["加班天数"], 3)
        self.assertEqual(detail["加班费"], 1050)

    def test_hanyuefu_june_uses_verified_overtime_public_rule(self):
        rows = calculate(
            [], [{"name": "程军虎", "project": "瀚阅府", "category": "益通外包",
                  "position_salary": 3800, "performance": 1000}],
            [{"name": "程军虎", "project": "瀚阅府", "category": "益通外包",
              "actual_work_days": 26, "overtime_days": 4}],
            "2026-06", 26, 4, "single",
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["加班天数"], 4)
        self.assertEqual(detail["加班费"], 1400)

    def test_june_holiday_reduces_base_workdays_before_overtime_cap(self):
        rows = calculate(
            [], [{"name": "节日", "project": "瀚阅府", "category": "益通外包",
                  "position_salary": 3800, "performance": 1000}],
            [{"name": "节日", "project": "瀚阅府", "category": "益通外包",
              "actual_work_days": 25, "holiday_dates": [
                  "2026-06-19", "2026-06-20", "2026-06-21",
              ]}],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["加班天数"], 4)

    def test_unique_ocr_name_correction_is_marked_for_review(self):
        rows = calculate(
            [], [{"name": "熊艳丽", "project": "瀚阅府", "category": "益通外包",
                  "position_salary": 3000}],
            [{"name": "熊艳利", "project": "瀚阅府", "category": "益通外包",
              "actual_work_days": 21}],
            "2026-06", 21, 4, "single",
        )
        self.assertEqual(rows["payroll_detail"][0]["项目"], "瀚阅府")
        self.assertTrue(any("OCR 近似" in item["说明"] for item in rows["review_exceptions"]))

    def test_name_ocr_correction_is_reported_once(self):
        rows = calculate(
            [{"name": "任新罡", "project": "瀚阅府", "category": "益通外包"}],
            [{"name": "任新罡", "project": "瀚阅府", "category": "益通外包",
              "position_salary": 3000}],
            [{"name": "任新翌", "project": "瀚阅府", "category": "益通外包",
              "actual_work_days": 21}],
            "2026-06", 21, 4,
        )
        notes = [
            item["说明"] for item in rows["review_exceptions"]
            if item["姓名"] == "任新翌" and "OCR 近似" in item["说明"]
        ]
        self.assertEqual(notes, ["姓名按 OCR 近似候选匹配，需人工复核"])

    def test_roster_project_has_priority_over_attendance_project(self):
        rows = calculate(
            [{"name": "甲", "project": "瀚阅府", "category": "外聘", "status": "正常"}],
            [{"name": "甲", "project": "瀚阅府", "category": "外聘",
              "position_salary": 3000, "performance": 500}],
            [{"name": "甲", "project": "瀚阅府西苑项目", "category": "外聘",
              "actual_work_days": 21}],
            "2026-06", 21, 4,
            attendance_project="瀚阅府西苑项目",
        )
        self.assertEqual(rows["payroll_detail"][0]["项目"], "瀚阅府")

    def test_job_title_misread_as_project_does_not_block_matching(self):
        rows = calculate(
            [{"name": "翁小龙", "project": "瀚阅府", "category": "易通外包", "status": "正常"}],
            [{"name": "翁小龙", "project": "瀚阅府", "category": "易通外包",
              "position_salary": 2300, "performance": 2500}],
            [{"name": "翁小龙", "project": "施工员", "actual_work_days": 26}],
            "2026-06", 21, 4, attendance_project="瀚阅府",
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["项目"], "瀚阅府")
        self.assertEqual(detail["岗位工资"], 2300)

    def test_unknown_vision_placeholder_is_not_a_payroll_person(self):
        rows = calculate(
            [], [{"name": "未知人员", "project": "瀚阅府", "category": "外聘",
                  "position_salary": 3000}],
            [{"name": "无法识别", "project": "施工员", "actual_work_days": 20}],
            "2026-06", 21, 4, attendance_project="瀚阅府",
        )
        self.assertEqual(rows["payroll_detail"], [])

    def test_ignored_review_always_uses_enrolled_category(self):
        rows = calculate(
            [{"name": "杨福生", "project": "瀚阅府", "category": "项目经理", "status": "在册"}],
            [], [{"name": "杨福生", "project": "瀚阅府", "category": "项目经理",
                   "actual_work_days": 21}],
            "2026-06", 21, 4,
        )
        review = rows["review_exceptions"][0]
        self.assertEqual(review["人员类别"], "在册")

    def test_department_label_does_not_block_employment_category_match(self):
        rows = calculate(
            [], [{"name": "李福山", "project": "瀚阅府", "category": "益通外包",
                  "position_salary": 3000, "performance": 500}],
            [{"name": "李福山", "project": "瀚阅府", "category": "项目管理人员",
              "actual_work_days": 21}],
            "2026-06", 21, 4, "single",
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["人员类别"], "益通外包")
        self.assertEqual(detail["岗位工资"], 3000)

    def test_project_suffix_and_missing_category_still_match_unique_person(self):
        rows = calculate(
            [{"name": "周九", "project": "瀚阅府", "category": "外聘", "status": "正常"}],
            [{"name": "周九", "project": "瀚阅府", "category": "外聘",
              "position_salary": 5000, "performance": 1000,
              "construction_day": 20}],
            [{"name": "周九", "project": "瀚阅府西苑项目",
              "actual_work_days": 20}],
            "2026-06", 22, 8,
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["岗位工资"], 5000)
        self.assertEqual(detail["施工补贴"], 400)
        self.assertEqual(detail["人员类别"], "外聘")
        self.assertEqual(detail["计算状态"], "需复核")

    def test_chinese_attendance_days_are_read(self):
        rows = calculate(
            [{"name": "吴十", "project": "G", "category": "外聘", "status": "正常"}],
            [{"name": "吴十", "project": "G", "category": "外聘",
              "position_salary": 5000, "performance": 1000, "construction_day": 20}],
            [{"name": "吴十", "project": "G", "实际出勤天数": 23}],
            "2026-06", 22, 8,
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["实际出勤"], 23)
        self.assertEqual(detail["施工补贴"], 460)

    def test_vision_actual_attendance_and_top_level_project_are_used(self):
        with tempfile.TemporaryDirectory() as directory:
            attendance_path = Path(directory) / "attendance.json"
            attendance_path.write_text(
                '{"month":"2026-06","project":"瀚阅府西苑项目","records":'
                '[{"name":"程军虎","actual_attendance":26}]}',
                encoding="utf-8",
            )
            month, project, attendance = attendance_records(attendance_path)
            rows = calculate(
                [{"name": "程军虎", "project": "瀚阅府", "category": "益通外包", "status": "正常"}],
                [{"name": "程军虎", "project": "瀚阅府", "category": "益通外包",
                  "position_salary": 5000, "performance": 1000, "construction_day": 20}],
                attendance, month, 22, 8, attendance_project=project,
            )
            detail = rows["payroll_detail"][0]
            review = rows["review_exceptions"][0]
            self.assertEqual(detail["实际出勤"], 26)
            self.assertEqual(detail["施工补贴"], 520)
            self.assertEqual(detail["项目"], "瀚阅府")
            self.assertEqual(detail["人员类别"], "益通外包")
            self.assertIn("人员类别", review)


if __name__ == "__main__":
    unittest.main()
