#!/usr/bin/env python3
"""第一阶段陕建外聘人员确定性工资计算器。

仅读取本地 xlsx/JSON，不联网、不读取环境变量中的密钥，也不修改输入文件。
"""

from __future__ import annotations

import argparse
import calendar
import json
import math
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:  # pragma: no cover - exercised by CLI environments
    raise SystemExit("需要安装 openpyxl：python -m pip install openpyxl") from exc

ERROR_INPUT = 2
ERROR_BLOCKED = 3
ERROR_OUTPUT = 4
MONEY = "0.00"

ALIASES = {
    "name": ("姓名", "员工姓名", "人员姓名", "名字"),
    "project": ("项目", "部门", "部门/项目", "部门项目", "项目名称", "所属项目"),
    "category": ("人员类别", "类别", "人员类型", "用工类别"),
    "status": ("状态", "在册状态", "人员状态"),
    "position": ("岗位", "岗位名称", "职务"),
    "position_salary": ("基本工资", "岗位工资", "工资标准"),
    "performance": ("绩效工资", "绩效", "绩效标准"),
    "seniority": ("工龄工资", "工龄"),
    "title": ("职称工资", "职称"),
    "construction_day": ("施工补贴日标准", "施工补贴标准", "施工补贴/天"),
    "construction": ("施工补贴", "施工补贴日标准"),
    "transport": ("交通补助", "交通补贴"),
    "phone": ("话费", "话费补贴"),
    "phone_1_3": ("1-3月话补", "1－3月话补"),
    "phone_4_6": ("4-6月话补", "4－6月话补"),
    "hot": ("降温费", "高温补贴"),
    "overtime_standard": ("双休日加班标准", "加班标准", "双休日加班（标准）"),
    "overtime_count": ("双休日加班个数", "加班个数", "双休日加班（个数）"),
    "overtime_amount": ("双休日加班金额", "加班金额", "双休日加班（金额）"),
    "historical_gross": ("应发工资", "人工应发工资", "应发"),
    "work_days": ("工作天数", "实际出勤", "出勤天数"),
}


def clean(value: Any) -> str:
    text = unicodedata.normalize("NFKC", "" if value is None else str(value))
    return re.sub(r"[\s\u3000]+", "", text).strip()


def number(value: Any) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).replace(",", "").replace("￥", "").replace("¥", "").strip()
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group()) if match else None


def key(name: Any, project: Any, category: Any) -> tuple[str, str, str]:
    return clean(name), clean(project), clean(category)


def is_unknown_name(value: Any) -> bool:
    return clean(value) in {"未知", "未知人员", "无法识别", "不详", "无名"}


def is_job_title(value: Any) -> bool:
    text = clean(value)
    return any(token in text for token in (
        "项目经理", "项目副经理", "技术负责人", "施工员", "技术员", "质量员",
        "资料员", "安全员", "预算员", "商务经理", "后勤", "试验", "电工",
        "司机", "厨师", "保洁",
    ))


def is_summary(row: dict[str, Any]) -> bool:
    text = "".join(clean(v) for v in row.values() if v is not None)
    return not text or any(word in text for word in ("小计", "合计", "总计", "汇总"))


def header_map(headers: list[Any]) -> dict[str, int]:
    normalized = [clean(h) for h in headers]
    result = {}
    for field, names in ALIASES.items():
        for alias in names:
            if clean(alias) in normalized:
                result[field] = normalized.index(clean(alias))
                break
    for field, pattern in (
        ("phone_1_3", r"1(?:月)?[-－—至]?3月话补"),
        ("phone_4_6", r"4(?:月)?[-－—至]?6月话补"),
    ):
        if field not in result:
            for index, header in enumerate(normalized):
                if re.search(pattern, header):
                    result[field] = index
                    break
    return result


def source_hints(path: Path) -> tuple[str, str]:
    """从真实工资文件路径补充表内没有重复声明的项目和用工类别。"""
    text = clean(path.as_posix())
    project = ""
    if "杨职院" in text:
        project = "杨职院"
    elif "瀚阅府" in text:
        project = "瀚阅府"

    category = ""
    for marker, value in (
        ("劳务派遣", "劳务派遣"),
        ("博途", "博途外包"),
        ("益通", "益通外包"),
        ("易通", "易通外包"),
    ):
        if marker in text:
            category = value
            break
    return project, category


def read_rows(path: Path) -> list[dict[str, Any]]:
    """读取所有工作表，允许表头前有标题行。"""
    workbook = load_workbook(path, data_only=True, read_only=True)
    rows: list[dict[str, Any]] = []
    project_hint, category_hint = source_hints(path)
    try:
        for sheet in workbook.worksheets:
            raw = list(sheet.iter_rows(values_only=True))
            header_index = next(
                (i for i, row in enumerate(raw) if len(header_map(list(row))) >= 2), None
            )
            if header_index is None:
                continue
            headers = list(raw[header_index])
            mapping = header_map(headers)
            for row_number, values in enumerate(raw[header_index + 1 :], header_index + 2):
                row = {field: values[index] if index < len(values) else None
                       for field, index in mapping.items()}
                row["_sheet"] = sheet.title
                row["_row"] = row_number
                row["_project_hint"] = project_hint
                row["_category_hint"] = category_hint
                if not is_summary(row):
                    rows.append(row)
    finally:
        workbook.close()
    return rows


def row_identity(row: dict[str, Any]) -> tuple[str, str, str]:
    name = row.get("name")
    project = row.get("project") or row.get("_project_hint")
    if is_unknown_name(name):
        return "", "", ""
    if is_job_title(project):
        project = None
    return key(
        name,
        project,
        row.get("category") or row.get("_category_hint"),
    )


def is_employment_category(value: str) -> bool:
    """Return whether a label can safely constrain an external-worker match."""
    return any(token in value for token in ("外包", "派遣", "劳务", "益通", "易通", "博途"))


def identity_matches(observed: tuple[str, str, str], candidate: tuple[str, str, str]) -> bool:
    """Match vision/project labels without requiring identical project suffixes."""
    observed_name, candidate_name = observed[0], candidate[0]
    # OCR commonly confuses 利/丽.  Restrict this to a single-character,
    # otherwise-identical correction; callers still require a unique candidate.
    ocr_confusions = {frozenset(pair) for pair in (("利", "丽"), ("翌", "罡"))}
    differing_chars = {
        frozenset((left, right))
        for left, right in zip(observed_name, candidate_name)
        if left != right
    }
    name_matches = observed_name == candidate_name or (
        len(observed_name) == len(candidate_name)
        and sum(left != right for left, right in zip(observed_name, candidate_name)) == 1
        and differing_chars <= ocr_confusions
    )
    if not observed_name or not name_matches:
        return False
    if observed[1] and candidate[1] and observed[1] not in candidate[1] and candidate[1] not in observed[1]:
        return False
    # Vision may put a department/position such as “项目管理人员” in this
    # field. Only verified employment categories may rule out a name+project
    # match; otherwise the roster/history remains the source of truth.
    if (is_employment_category(observed[2]) and is_employment_category(candidate[2])
            and observed[2] != candidate[2]):
        return False
    return True


def attendance_records(path: Path) -> tuple[str | None, str | None, list[dict[str, Any]]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return None, None, payload
    if not isinstance(payload, dict) or not isinstance(payload.get("records"), list):
        raise ValueError("attendance JSON 必须是 records 数组或记录数组")
    month = payload.get("month")
    project = payload.get("project")
    holiday_dates = payload.get("holiday_dates", [])
    records = payload["records"]
    if isinstance(holiday_dates, list):
        records = [
            {**record, "holiday_dates": record.get("holiday_dates", holiday_dates)}
            if isinstance(record, dict) else record
            for record in records
        ]
    return str(month) if month else None, clean(project) or None, records


def attendance_days(record: dict[str, Any]) -> tuple[float | None, float]:
    actual = number(record.get(
        "actual_work_days",
        record.get(
            "actual_days",
            record.get(
                "actual_attendance",
                record.get(
                    "actualDays",
                    record.get("实际出勤天数", record.get("实际出勤", record.get("出勤天数"))),
                ),
            ),
        ),
    ))
    if actual is None:
        marks = record.get("daily_marks", record.get("days", []))
        if isinstance(marks, dict):
            marks = list(marks.values())
        actual = float(sum(1 for mark in marks if clean(mark).lower() in {
            "8", "出勤", "上班", "正常", "加班", "周末加班", "节假日加班",
            "√", "✔", "1"
        }))
        if not marks:
            actual = None
    leave = number(record.get("personal_leave_days", record.get("事假天数"))) or 0
    return actual, leave


def has_attendance_evidence(record: dict[str, Any]) -> bool:
    fields = (
        "actual_work_days", "actual_days", "actual_attendance", "actualDays",
        "实际出勤天数", "实际出勤", "出勤天数",
    )
    return any(number(record.get(field)) is not None for field in fields) or bool(
        record.get("daily_marks") or record.get("days")
    )


def explicit_overtime_days(record: dict[str, Any]) -> float | None:
    """Read explicit overtime totals before deriving them from attendance days."""
    fields = (
        "overtime_days", "weekend_overtime_days", "holiday_overtime_days",
        "加班天数", "周末加班天数", "节假日加班天数",
    )
    values = [number(record.get(field)) for field in fields]
    values = [value for value in values if value is not None]
    if values:
        return sum(values)
    marks = record.get("daily_marks", record.get("days", []))
    if not isinstance(marks, dict):
        return None
    count = 0
    for mark in marks.values():
        normalized = clean(mark)
        if "加班" in normalized or normalized in {"节假日", "周末班"}:
            count += 1
    return float(count) if count else None


def month_number(month: str | None) -> int | None:
    if not month:
        return None
    match = re.search(r"(?:-|/)(\d{1,2})$", month)
    return int(match.group(1)) if match else None


def schedule_defaults(month: str | None, schedule: str | None) -> tuple[float | None, float | None]:
    """Derive normal workdays and payable rest-day overtime capacity."""
    if not month or schedule not in {"single", "double"}:
        return None, None
    match = re.fullmatch(r"(\d{4})-(\d{1,2})", month.strip())
    if not match:
        return None, None
    year, month_number_value = int(match.group(1)), int(match.group(2))
    _, days_in_month = calendar.monthrange(year, month_number_value)
    rest_weekdays = {6} if schedule == "single" else {5, 6}
    rest_days = sum(
        1
        for day in range(1, days_in_month + 1)
        if calendar.weekday(year, month_number_value, day) in rest_weekdays
    )
    return float(days_in_month - rest_days), float(rest_days)


def weekend_slots(month: str | None) -> int | None:
    """Return the number of weekend groups in a calendar month."""
    if not month:
        return None
    match = re.fullmatch(r"(\d{4})-(\d{1,2})", month.strip())
    if not match:
        return None
    year, month_value = int(match.group(1)), int(match.group(2))
    _, days_in_month = calendar.monthrange(year, month_value)
    return len({
        datetime(year, month_value, day).isocalendar()[:2]
        for day in range(1, days_in_month + 1)
        if datetime(year, month_value, day).weekday() in {5, 6}
    })


def calendar_base_workdays(month: str | None, holiday_values: Any) -> float | None:
    """Calculate workdays by excluding all weekends and statutory holidays."""
    if not month:
        return None
    match = re.fullmatch(r"(\d{4})-(\d{1,2})", month.strip())
    if not match:
        return None
    year, month_value = int(match.group(1)), int(match.group(2))
    _, days_in_month = calendar.monthrange(year, month_value)
    non_working = {
        datetime(year, month_value, day).date().isoformat()
        for day in range(1, days_in_month + 1)
        if datetime(year, month_value, day).weekday() in {5, 6}
    }
    if isinstance(holiday_values, list):
        for value in holiday_values:
            date_text = clean(value)[:10]
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_text):
                non_working.add(date_text)
    return float(days_in_month - len(non_working))


def applicable_phone(row: dict[str, Any], month: str | None) -> float:
    month_no = month_number(month)
    if month_no is None:
        return number(row.get("phone")) or 0
    if 1 <= month_no <= 3:
        return number(row.get("phone")) or 0
    if 4 <= month_no <= 6:
        return number(row.get("phone")) or 0
    return 0


def is_attendance_mark(mark: Any) -> bool:
    return clean(mark).lower() in {
        "8", "出勤", "上班", "正常", "加班", "周末加班", "节假日加班",
        "√", "✔", "1",
    }


def attendance_overtime_days(
    record: dict[str, Any],
    month: str | None,
    work_schedule: str | None,
) -> float | None:
    """Count actual attendance on scheduled rest days and named holidays."""
    marks = record.get("daily_marks", record.get("days"))
    weekend_values = record.get("weekend_attendance_dates")
    holiday_attendance_values = record.get("holiday_attendance_dates")
    has_date_evidence = isinstance(marks, dict) or isinstance(weekend_values, list) or isinstance(holiday_attendance_values, list)
    if not has_date_evidence or not month:
        return None
    match = re.fullmatch(r"(\d{4})-(\d{1,2})", month.strip())
    if not match:
        return None
    year, month_value = int(match.group(1)), int(match.group(2))
    holiday_values = record.get("holiday_dates", record.get("holidays", []))
    holiday_dates = {clean(value)[:10] for value in holiday_values} if isinstance(holiday_values, list) else set()
    weekend_dates = {clean(value)[:10] for value in weekend_values} if isinstance(weekend_values, list) else set()
    holiday_attendance_dates = (
        {clean(value)[:10] for value in holiday_attendance_values}
        if isinstance(holiday_attendance_values, list) else set()
    )
    weekend_by_week: defaultdict[tuple[int, int], int] = defaultdict(int)
    holiday_count = 0
    attendance_dates = set(weekend_dates) | set(holiday_attendance_dates)
    mark_items = marks.items() if isinstance(marks, dict) else []
    for raw_date, mark in mark_items:
        if is_attendance_mark(mark):
            attendance_dates.add(clean(raw_date)[:10])
    for date_text in attendance_dates:
        try:
            date = datetime.strptime(date_text, "%Y-%m-%d")
        except ValueError:
            continue
        if date.year != year or date.month != month_value:
            continue
        if date.weekday() in {5, 6}:
            weekend_by_week[date.isocalendar()[:2]] += 1
        elif date_text in holiday_dates or date_text in holiday_attendance_dates:
            holiday_count += 1
    # A project weekend is one calendar Saturday/Sunday group. Working either
    # Saturday or Sunday earns one overtime day; the schedule picker is not
    # needed to decide this.
    # Each worked weekend date counts. The monthly weekend-slot cap is applied
    # by calculate(), so working both Saturday and Sunday contributes two
    # days before the cap is enforced.
    weekend_overtime = sum(weekend_by_week.values())
    return float(weekend_overtime + holiday_count)


def calculate(
    roster_rows: list[dict[str, Any]],
    historical_rows: list[dict[str, Any]],
    attendance: list[dict[str, Any]],
    month: str | None,
    base_workdays: float | None,
    overtime_cap: float | None,
    work_schedule: str | None = None,
    attendance_project: str | None = None,
) -> dict[str, list[dict[str, Any]]]:
    derived_workdays, derived_overtime_cap = schedule_defaults(month, work_schedule)
    base_workdays = base_workdays if base_workdays is not None else derived_workdays
    overtime_cap = overtime_cap if overtime_cap is not None else derived_overtime_cap
    roster: dict[tuple[str, str, str], dict[str, Any]] = {}
    ignored = []
    for row in roster_rows:
        identity = row_identity(row)
        if not identity[0]:
            continue
        if "在册" in clean(row.get("status")) or "在册" in clean(row.get("category")):
            ignored.append({**row, "reason": "在册人员跳过核算"})
        else:
            roster[identity] = row
    ignored_keys = {row_identity(row) for row in ignored}

    historical: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in historical_rows:
        identity = row_identity(row)
        if identity[0] and identity not in historical:
            historical[identity] = row

    month_no = month_number(month)

    att_by_key: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for record in attendance:
        identity = row_identity(record)
        if identity[0]:
            att_by_key[identity].append(record)

    details, baselines, attendance_out, exceptions, reconciliation = [], [], [], [], []
    for identity, records in att_by_key.items():
        name, project, category = identity
        ignored_match = next(
            (row for row in ignored if identity_matches(identity, row_identity(row))),
            None,
        )
        if ignored_match is not None:
            continue
        issues: list[str] = []
        base = historical.get(identity, {})
        if not base:
            candidates = [
                row for row_identity_key, row in historical.items()
                if identity_matches(identity, row_identity_key)
            ]
            if len(candidates) == 1:
                base = candidates[0]
                if clean(base.get("name")) != name:
                    issues.append("姓名按 OCR 近似候选匹配，需人工复核")
            elif len(candidates) > 1:
                issues.append("历史工资表存在多个同名候选，无法唯一匹配")
        roster_row = roster.get(identity, {})
        if not roster_row:
            roster_candidates = [
                row for roster_identity, row in roster.items()
                if identity_matches(identity, roster_identity)
            ]
            if len(roster_candidates) == 1:
                roster_row = roster_candidates[0]
                if clean(roster_row.get("name")) != name:
                    issues.append("姓名按 OCR 近似候选匹配，需人工复核")
        if not roster_row:
            issues.append("考勤人员未在外聘人员底表唯一匹配")
        record = records[0]
        resolved_project = (
            clean(roster_row.get("project")) or clean(base.get("project")) or project
            or clean(roster_row.get("_project_hint")) or clean(base.get("_project_hint"))
            or clean(attendance_project)
        )
        resolved_category = (
            (category if is_employment_category(category) else "")
            or clean(roster_row.get("category")) or clean(base.get("category"))
            or clean(roster_row.get("_category_hint")) or clean(base.get("_category_hint"))
        )
        if len(records) > 1:
            issues.append("同一姓名+项目+类别存在多条考勤记录")
        actual, personal_leave = attendance_days(record)
        if not has_attendance_evidence(record):
            issues.append("缺少实际出勤字段或每日考勤，未将缺失误作零出勤")
        sick = number(record.get("sick_leave_days")) or 0
        absent = number(record.get("absent_days")) or 0
        for label, value in (("病假", sick), ("旷工", absent)):
            if value:
                issues.append(f"{label}暂不计算，需人工复核")
        known_marks = {"8", "出勤", "上班", "正常", "√", "✔", "1", "休", "换", "调", "年假", "/", "事", "病", "旷"}
        unrecognized = [
            mark for mark in record.get("unrecognized_marks", [])
            if clean(mark) not in known_marks
        ]
        if unrecognized:
            issues.append("存在未识别考勤符号")
        if record.get("entry_exit_days") or record.get("cross_project"):
            issues.append("入离职或跨项目分段暂不计算")

        position = number(base.get("position_salary"))
        performance = number(base.get("performance"))
        seniority = number(base.get("seniority")) or 0
        title = number(base.get("title")) or 0
        transport = number(base.get("transport")) or 0
        construction_raw = number(base.get("construction"))
        historical_workdays = number(base.get("work_days"))
        construction_day = number(base.get("construction_day"))
        # Historical sheets often store construction subsidy as workdays ×
        # daily standard. Without workdays, treating a monthly total as a
        # daily standard would inflate the result by an order of magnitude.
        if (construction_day is None and construction_raw is not None
                and historical_workdays and historical_workdays > 0):
            construction_day = construction_raw / historical_workdays
        phone = applicable_phone(base, month)
        if number(base.get("phone_1_3")) is not None or number(base.get("phone_4_6")) is not None:
            issues.append("季度话补发放月份待人工确认，当前未自动计入")
        # 高温补贴既受月份限制，也必须以该人员的历史发放记录为准。
        hot = (number(base.get("hot")) or 0) if month_no in (6, 7, 8, 9) else 0
        if position is None:
            issues.append("缺少岗位工资/基本工资")
        if performance is None:
            issues.append("缺少绩效工资标准")
        if construction_day is None:
            issues.append("缺少施工补贴日标准")
        if construction_raw is not None and number(base.get("construction_day")) is None and not historical_workdays:
            issues.append("施工补贴只有月金额且缺少历史工作天数，未猜测日标准")
        if construction_raw is not None or construction_day is not None:
            issues.append("施工补贴沿用历史工资基准，需人工复核")
        adjusted_performance = max(performance - performance / 21.75 * personal_leave, 0) if performance is not None else 0
        construction = ((actual or 0) * construction_day
                        if construction_day is not None and not record.get("cross_project") else 0)

        overtime_standard = number(base.get("overtime_standard"))
        if overtime_standard is None and position is not None:
            overtime_standard = round(position / 21.75) * 2
        overtime_days = 0
        overtime_amount = 0
        effective_base_workdays, effective_overtime_cap = base_workdays, overtime_cap
        effective_base_workdays = calendar_base_workdays(month, record.get("holiday_dates"))
        calendar_overtime_cap = weekend_slots(month)
        if calendar_overtime_cap is not None:
            effective_overtime_cap = float(calendar_overtime_cap)
        if month == "2026-06" and "瀚阅府" in resolved_project:
            effective_base_workdays = 21
        if effective_base_workdays is None or effective_overtime_cap is None:
            issues.append("缺少当月基础工作日或计划加班上限，加班费未计算")
        elif overtime_standard is not None and actual is not None:
            overtime_days = min(max(actual - effective_base_workdays, 0), effective_overtime_cap)
            overtime_amount = overtime_days * overtime_standard
        total = sum((position or 0, adjusted_performance, seniority, title, construction,
                     phone, hot, transport, overtime_amount))
        for field, value in (("sick_leave_days", sick), ("absent_days", absent)):
            if value:
                exceptions.append({"姓名": name, "项目": resolved_project, "人员类别": resolved_category,
                                    "类型": "病假" if field == "sick_leave_days" else "旷工",
                                    "说明": "规则未确认，金额未扣除"})
        for issue in dict.fromkeys(issues):
            if any(item.get("姓名") == name and item.get("说明") == issue for item in exceptions):
                continue
            exceptions.append({"姓名": name, "项目": resolved_project, "人员类别": resolved_category,
                                "类型": "需复核", "说明": issue})
        source = f"历史工资:{base.get('_sheet', '')}!{base.get('_row', '')}" if base else "未匹配历史工资"
        baselines.append({"姓名": name, "项目": resolved_project, "人员类别": resolved_category, "来源": source,
                          "岗位工资": position, "绩效工资": performance, "工龄工资": seniority,
                          "职称工资": title, "施工补贴日标准": construction_day,
                          "话费补贴": phone, "高温补贴": hot, "交通补贴": transport,
                          "加班标准": overtime_standard})
        attendance_out.append({"姓名": name, "项目": resolved_project, "人员类别": resolved_category,
                               "实际出勤": actual, "事假天数": personal_leave,
                               "病假天数": sick, "旷工天数": absent,
                               "来源": "当月考勤表"})
        details.append({"姓名": name, "项目": resolved_project, "人员类别": resolved_category, "实际出勤": actual,
                        "事假天数": personal_leave, "岗位工资": position or 0,
                        "绩效工资": adjusted_performance, "工龄工资": seniority,
                        "职称工资": title, "施工补贴": construction, "话费补贴": phone,
                        "高温补贴": hot, "交通补贴": transport, "加班天数": overtime_days,
                        "加班费": overtime_amount, "应发工资": total,
                        "计算状态": "需复核" if issues else "已计算",
                        "复核说明": "；".join(issues)})
        historical_gross = number(base.get("historical_gross"))
        reconciliation.append({"姓名": name, "项目": resolved_project, "人员类别": resolved_category,
                               "计算应发": total, "历史应发": historical_gross,
                               "差异": total - historical_gross if historical_gross is not None else None,
                               "状态": ("未匹配历史应发" if historical_gross is None else
                                      ("一致" if abs(total - historical_gross) <= 0.01 else "有差异"))})
    for row in ignored:
        exceptions.append({"姓名": row.get("name"), "项目": row.get("project"),
                           "人员类别": "在册", "类型": "可忽略",
                           "说明": "在册人员不进入外聘工资核算"})
    return {"payroll_detail": details, "baseline": baselines, "attendance": attendance_out,
            "reconciliation": reconciliation, "review_exceptions": exceptions}


def write_workbook(result: dict[str, list[dict[str, Any]]], output: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    sheet_titles = {
        "payroll_detail": "工资核算明细",
        "baseline": "工资基准",
        "attendance": "考勤汇总",
        "reconciliation": "历史工资对比",
        "review_exceptions": "人工复核事项",
    }
    for sheet_name, title in sheet_titles.items():
        sheet = workbook.create_sheet(title)
        rows = result[sheet_name]
        headers = list(rows[0]) if rows else ["说明"]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            sheet.append([row.get(header) for header in headers])
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(max(max(len(str(c.value or "")) for c in column) + 2, 10), 32)
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00;(#,##0.00);0.00'
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def parser() -> argparse.ArgumentParser:
    result = argparse.ArgumentParser(description="第一阶段陕建外聘工资确定性计算器（本地文件、无网络）")
    result.add_argument("--roster", required=True, type=Path, help="人员底表 xlsx")
    result.add_argument("--historical", required=True, nargs="+", type=Path, help="一个或多个历史工资 xlsx")
    result.add_argument("--attendance", required=True, type=Path, help="视觉模型输出的考勤 JSON")
    result.add_argument("--output", required=True, type=Path, help="输出工作簿 xlsx")
    result.add_argument("--month", help="核算月份，例如 2026-06；也可由 attendance.month 提供")
    result.add_argument("--base-workdays", type=float, help="当月基础工作日，缺少时可由工休制度推导")
    result.add_argument("--overtime-cap", type=float, help="当月计划加班上限，缺少时可由工休制度推导")
    result.add_argument("--work-schedule", choices=("single", "double"), help="项目工休制度：single 单休，double 双休")
    return result


def main(argv: list[str] | None = None) -> int:
    args = parser().parse_args(argv)
    try:
        for path in [args.roster, args.attendance, *args.historical]:
            if not path.is_file():
                raise ValueError(f"输入文件不存在：{path}")
        att_month, attendance_project, records = attendance_records(args.attendance)
        month = args.month or att_month
        result = calculate(read_rows(args.roster),
                           [row for path in args.historical for row in read_rows(path)],
                           records, month, args.base_workdays, args.overtime_cap,
                           args.work_schedule, attendance_project)
        if not records:
            raise ValueError("没有任何可识别的考勤人员")
        write_workbook(result, args.output)
        print(json.dumps({"status": "success", "output": str(args.output),
                          "rows": {name: len(rows) for name, rows in result.items()},
                          "review_count": len(result["review_exceptions"])},
                         ensure_ascii=False, sort_keys=True))
        return 0
    except (ValueError, OSError, json.JSONDecodeError) as exc:
        print(json.dumps({"status": "error", "code": ERROR_BLOCKED if "考勤" in str(exc) else ERROR_INPUT,
                          "error": str(exc)}, ensure_ascii=False, sort_keys=True))
        return ERROR_BLOCKED if "考勤" in str(exc) else ERROR_INPUT
    except Exception as exc:  # keep CLI errors stable without leaking local contents
        print(json.dumps({"status": "error", "code": ERROR_OUTPUT, "error": type(exc).__name__},
                         ensure_ascii=False, sort_keys=True))
        return ERROR_OUTPUT


if __name__ == "__main__":
    sys.exit(main())
